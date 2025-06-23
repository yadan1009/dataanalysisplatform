from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
import requests
import os
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
import json
import PyPDF2
import docx
import chardet
from datetime import datetime
import uuid
import pandas as pd
from openpyxl import load_workbook
import re
import traceback
import pypinyin  # 添加pypinyin库
from pypinyin import lazy_pinyin  # 直接导入lazy_pinyin函数
import matplotlib
matplotlib.use('Agg')  # 使用非交互式后端
import matplotlib.pyplot as plt
import numpy as np
import io
import base64
from matplotlib.font_manager import FontProperties
import seaborn as sns
from PIL import Image

# 配置中文字体
try:
    # 尝试加载中文字体
    font_path = '/System/Library/Fonts/PingFang.ttc'  # macOS中文字体路径
    if os.path.exists(font_path):
        plt.rcParams['font.family'] = ['PingFang SC', 'sans-serif']
        font = FontProperties(fname=font_path)
    else:
        # 如果没有PingFang字体，尝试其他常见中文字体
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'STHeiti', 'Arial Unicode MS', 'sans-serif']
        # 在matplotlib内部配置中添加以下设置
        plt.rc('font', family='sans-serif')
    
    # 全局设置默认字体
    matplotlib.rcParams['font.sans-serif'] = ['SimHei', 'PingFang SC', 'Microsoft YaHei', 'STHeiti', 'Arial Unicode MS', 'sans-serif']
    matplotlib.rcParams['axes.titlesize'] = 14
    matplotlib.rcParams['axes.labelsize'] = 12
except Exception as e:
    print(f"字体加载错误: {str(e)}")

plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

# 加载环境变量
load_dotenv()

app = Flask(__name__)
CORS(app)  # 允许跨域请求

# 阿里云大模型 API 配置
API_KEY = "YOUR_API_KEY"
API_URL = "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation"

# 文件上传配置
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'doc', 'docx', 'md', 'xlsx', 'xls'}
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
    
# Excel处理文件夹 - 使用绝对路径
PROCESSED_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'processed')
if not os.path.exists(PROCESSED_FOLDER):
    os.makedirs(PROCESSED_FOLDER)

# 数据分析结果文件夹
ANALYSIS_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'analysis')
if not os.path.exists(ANALYSIS_FOLDER):
    os.makedirs(ANALYSIS_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# 存储知识库内容和初始prompt
knowledge_base = {}
initial_prompt = {}
# 存储会话数据
sessions = {}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def detect_encoding(filepath):
    """检测文件编码"""
    with open(filepath, 'rb') as f:
        raw_data = f.read(10000)  # 读取前10000字节来检测编码
        result = chardet.detect(raw_data)
        return result['encoding']

def read_pdf_content(filepath):
    """读取PDF文件内容"""
    try:
        content = []
        with open(filepath, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            num_pages = len(pdf_reader.pages)
            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                content.append(page.extract_text())
        return '\n'.join(content)
    except Exception as e:
        return f"PDF读取错误: {str(e)}"

def read_docx_content(filepath):
    """读取DOCX文件内容"""
    try:
        doc = docx.Document(filepath)
        content = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                content.append(paragraph.text)
        return '\n'.join(content)
    except Exception as e:
        return f"DOCX读取错误: {str(e)}"

def read_file_content(filepath):
    """读取文件内容，支持多种格式"""
    try:
        # 获取文件扩展名
        file_extension = filepath.rsplit('.', 1)[1].lower() if '.' in filepath else ''
        
        # 根据文件类型选择读取方法
        if file_extension == 'pdf':
            return read_pdf_content(filepath)
        elif file_extension == 'docx':
            return read_docx_content(filepath)
        elif file_extension == 'doc':
            # DOC格式较复杂，建议转换为DOCX
            return "DOC格式文件暂不支持，请转换为DOCX格式后上传"
        else:
            # 文本文件：尝试检测编码
            encoding = detect_encoding(filepath)
            if not encoding:
                encoding = 'utf-8'
            
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    return f.read()
            except:
                # 如果检测的编码失败，尝试常用编码
                for enc in ['utf-8', 'gbk', 'gb2312', 'utf-16', 'big5']:
                    try:
                        with open(filepath, 'r', encoding=enc) as f:
                            return f.read()
                    except:
                        continue
                return "无法读取文件内容：编码格式不支持"
    except Exception as e:
        return f"文件读取错误: {str(e)}"

def create_session(mode='normal', title=None):
    """创建新会话"""
    session_id = str(uuid.uuid4())
    now = datetime.now()
    sessions[session_id] = {
        'id': session_id,
        'mode': mode,
        'title': title or f"新对话",
        'messages': [],
        'created_at': now.isoformat(),
        'updated_at': now.isoformat()
    }
    return sessions[session_id]

def update_session_title(session_id, first_message):
    """根据第一条消息更新会话标题"""
    if session_id in sessions and sessions[session_id]['title'].startswith("新对话"):
        # 截取前30个字符作为标题
        title = first_message[:30] + "..." if len(first_message) > 30 else first_message
        sessions[session_id]['title'] = title

@app.route('/api/chat', methods=['POST'])
def chat():
    try:
        # 获取前端发送的消息和模式
        data = request.get_json()
        user_message = data.get('message', '')
        mode = data.get('mode', 'normal')
        session_id = data.get('session_id', 'default')
        
        if not user_message:
            return jsonify({'error': '消息不能为空'}), 400
        
        # 如果会话不存在或者是新会话，创建新会话
        if session_id == 'new' or session_id not in sessions:
            session = create_session(mode)
            session_id = session['id']
            
            # 迁移临时设置到新会话
            if 'temp' in initial_prompt:
                initial_prompt[session_id] = initial_prompt['temp']
                del initial_prompt['temp']
            if 'temp' in knowledge_base:
                knowledge_base[session_id] = knowledge_base['temp']
                del knowledge_base['temp']
        else:
            session = sessions[session_id]
        
        # 如果是第一条消息，更新标题
        if len(session['messages']) == 0:
            update_session_title(session_id, user_message)
        
        # 构造请求头
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {API_KEY}"
        }
        
        # 构造消息列表
        messages = []
        
        # 根据模式添加系统消息
        if mode == 'normal' and session_id in initial_prompt:
            # 普通模式，使用初始prompt
            messages.append({
                "role": "system",
                "content": initial_prompt[session_id]
            })
        elif mode == 'knowledge' and session_id in knowledge_base:
            # 知识库模式，将知识库内容作为上下文
            kb_content = "\n\n".join(knowledge_base[session_id])
            system_message = f"你是一个智能助手。请基于以下知识库内容回答用户的问题。如果问题与知识库内容相关，请优先使用知识库中的信息进行回答。\n\n知识库内容：\n{kb_content}"
            messages.append({
                "role": "system",
                "content": system_message
            })
        
        # 添加历史消息（最多保留最近10轮对话）
        history_messages = session['messages'][-20:] if len(session['messages']) > 20 else session['messages']
        for msg in history_messages:
            messages.append({
                "role": msg['role'],
                "content": msg['content']
            })
        
        # 添加用户消息
        messages.append({
            "role": "user",
            "content": user_message
        })
        
        # 构造请求体
        payload = {
            "model": "qwen-max",
            "input": {
                "messages": messages
            },
            "parameters": {
                "temperature": 0.7,
                "max_tokens": 1000
            }
        }
        
        # 调用阿里云大模型 API
        response = requests.post(API_URL, headers=headers, json=payload)
        
        if response.status_code == 200:
            result = response.json()
            # 提取 AI 回复内容
            ai_reply = result.get('output', {}).get('text', '抱歉，我无法回答这个问题。')
            
            # 保存消息到会话历史
            session['messages'].append({
                'role': 'user',
                'content': user_message,
                'timestamp': datetime.now().isoformat()
            })
            session['messages'].append({
                'role': 'assistant',
                'content': ai_reply,
                'timestamp': datetime.now().isoformat()
            })
            session['updated_at'] = datetime.now().isoformat()
            
            return jsonify({
                'reply': ai_reply,
                'session_id': session_id
            })
        else:
            return jsonify({'error': 'API 调用失败'}), 500
            
    except Exception as e:
        return jsonify({'error': f'服务器错误: {str(e)}'}), 500

@app.route('/api/sessions', methods=['GET'])
def get_sessions():
    """获取所有会话列表"""
    try:
        # 只返回有消息的会话，按更新时间倒序排列
        session_list = [s for s in sessions.values() if len(s['messages']) > 0]
        session_list = sorted(session_list, 
                            key=lambda x: x['updated_at'], 
                            reverse=True)
        return jsonify(session_list)
    except Exception as e:
        return jsonify({'error': f'获取会话列表失败: {str(e)}'}), 500

@app.route('/api/sessions/<session_id>', methods=['GET'])
def get_session(session_id):
    """获取特定会话的详情"""
    try:
        if session_id in sessions:
            return jsonify(sessions[session_id])
        else:
            return jsonify({'error': '会话不存在'}), 404
    except Exception as e:
        return jsonify({'error': f'获取会话失败: {str(e)}'}), 500

@app.route('/api/sessions/<session_id>', methods=['DELETE'])
def delete_session(session_id):
    """删除会话"""
    try:
        if session_id in sessions:
            del sessions[session_id]
            # 同时删除相关的知识库和prompt
            if session_id in knowledge_base:
                del knowledge_base[session_id]
            if session_id in initial_prompt:
                del initial_prompt[session_id]
            return jsonify({'success': True})
        else:
            return jsonify({'error': '会话不存在'}), 404
    except Exception as e:
        return jsonify({'error': f'删除会话失败: {str(e)}'}), 500

@app.route('/api/sessions/new', methods=['POST'])
def new_session():
    """创建新会话"""
    try:
        data = request.get_json()
        mode = data.get('mode', 'normal')
        title = data.get('title', None)
        session = create_session(mode, title)
        return jsonify(session)
    except Exception as e:
        return jsonify({'error': f'创建会话失败: {str(e)}'}), 500

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """处理文件上传"""
    try:
        session_id = request.form.get('session_id', 'default')
        
        if 'file' not in request.files:
            return jsonify({'error': '没有文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
            file.save(filepath)
            
            # 读取文件内容并添加到知识库
            content = read_file_content(filepath)
            if session_id not in knowledge_base:
                knowledge_base[session_id] = []
            
            # 检查内容是否读取成功
            if "错误" in content or "不支持" in content:
                # 删除临时文件
                os.remove(filepath)
                return jsonify({'error': f'文件读取失败: {content}'}), 400
            
            # 添加到知识库，包含文件信息和内容长度
            file_info = f"文件：{filename}\n内容长度：{len(content)}字符\n内容：{content[:500]}..." if len(content) > 500 else f"文件：{filename}\n内容：{content}"
            knowledge_base[session_id].append(file_info)
            
            # 删除临时文件
            os.remove(filepath)
            
            return jsonify({'success': True, 'filename': filename, 'content_length': len(content)})
        else:
            return jsonify({'error': '不支持的文件类型'}), 400
            
    except Exception as e:
        return jsonify({'error': f'上传失败: {str(e)}'}), 500

@app.route('/api/set_prompt', methods=['POST'])
def set_prompt():
    """设置初始prompt"""
    try:
        data = request.get_json()
        session_id = data.get('session_id', 'default')
        prompt = data.get('prompt', '')
        
        initial_prompt[session_id] = prompt
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'error': f'设置失败: {str(e)}'}), 500

@app.route('/api/clear_knowledge', methods=['POST'])
def clear_knowledge():
    """清空知识库"""
    try:
        data = request.get_json()
        session_id = data.get('session_id', 'default')
        
        if session_id in knowledge_base:
            knowledge_base[session_id] = []
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'error': f'清空失败: {str(e)}'}), 500

@app.route('/api/process_excel', methods=['POST'])
def process_excel():
    if 'file' not in request.files:
        return jsonify({'error': '没有文件上传'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        temp_filepath = None
        unmerged_filepath = None
        try:
            filename = secure_filename(file.filename)
            temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(temp_filepath)

            # 1. 使用 openpyxl 处理合并单元格
            wb = load_workbook(temp_filepath)
            sheet = wb.active
            
            merged_cells_ranges = list(sheet.merged_cells.ranges)
            unmerged_count = 0
            for merged_range in merged_cells_ranges:
                unmerged_count += 1
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                sheet.unmerge_cells(str(merged_range))
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        sheet.cell(row=row, column=col).value = top_left_cell_value
            
            unmerged_filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'unmerged_' + filename)
            wb.save(unmerged_filepath)

            # 2. 使用 pandas 读取数据并处理复杂表头
            # 这是一个启发式方法，它假定标题行比数据行包含更多的空值
            df_temp = pd.read_excel(unmerged_filepath, header=None)
            header_rows = 0
            for i, row in df_temp.iterrows():
                if (row.isnull().sum() / len(row)) < 0.4:
                    break
                header_rows = i + 1
            
            if header_rows > 0:
                df = pd.read_excel(unmerged_filepath, header=list(range(header_rows)))
                df.columns = df.columns.map(lambda x: ' '.join(map(str, x)).strip().replace('Unnamed: 0', '').replace('nan', ''))
            else:
                df = pd.read_excel(unmerged_filepath)

            df.dropna(how='all', inplace=True)
            df.reset_index(drop=True, inplace=True)

            # 保存处理后的文件
            processed_filename = 'processed_' + str(uuid.uuid4()) + '.xlsx'
            processed_filepath = os.path.join(PROCESSED_FOLDER, processed_filename)
            df.to_excel(processed_filepath, index=False)
            
            # 清理临时文件
            os.remove(temp_filepath)
            os.remove(unmerged_filepath)

            # 生成预览数据
            preview_data = df.head(10).to_json(orient='split')
            
            summary = {
                "unmerged_cells": unmerged_count,
                "header_rows_detected": header_rows,
                "total_rows": len(df),
                "total_cols": len(df.columns)
            }

            return jsonify({
                'success': True,
                'preview': preview_data,
                'download_filename': processed_filename,
                'summary': summary
            })

        except Exception as e:
            if temp_filepath and os.path.exists(temp_filepath):
                os.remove(temp_filepath)
            if unmerged_filepath and os.path.exists(unmerged_filepath):
                os.remove(unmerged_filepath)
            return jsonify({'error': f'处理文件时发生错误: {str(e)}'}), 500
    else:
        return jsonify({'error': '不支持的文件类型，请上传 .xls 或 .xlsx 文件'}), 400

@app.route('/api/download_processed/<filename>')
def download_processed_file(filename):
    filepath = os.path.join(PROCESSED_FOLDER, filename)
    try:
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True)
        else:
            # 返回详细的路径信息和错误
            return jsonify({
                'error': '文件未找到',
                'filepath': filepath,
                'processed_folder_exists': os.path.exists(PROCESSED_FOLDER),
                'processed_folder': PROCESSED_FOLDER
            }), 404
    except Exception as e:
        return jsonify({
            'error': f'下载文件失败: {str(e)}',
            'filepath': filepath,
            'processed_folder': PROCESSED_FOLDER
        }), 500

@app.route('/api/excel_ai_process', methods=['POST'])
def excel_ai_process():
    try:
        data = request.get_json()
        instruction = data.get('instruction', '')
        preview_data = data.get('preview_data', {})
        filename = data.get('filename', '')
        
        if not instruction or not preview_data or not filename:
            return jsonify({'error': '缺少必要参数'}), 400
        
        # 获取Excel数据
        df = None
        try:
            # 首先尝试加载完整数据
            full_filepath = os.path.join(PROCESSED_FOLDER, filename)
            if os.path.exists(full_filepath):
                try:
                    full_df = pd.read_excel(full_filepath)
                    print(f"已加载完整数据，共 {len(full_df)} 行")
                    df = full_df
                except Exception as e:
                    print(f"加载完整数据失败: {str(e)}，将使用预览数据")
                    
            # 如果无法加载完整数据，使用预览数据
            if df is None:
                preview = json.loads(preview_data) if isinstance(preview_data, str) else preview_data
                columns = preview.get('columns', [])
                data_rows = preview.get('data', [])
                df = pd.DataFrame(data_rows, columns=columns)
                print(f"使用预览数据，共 {len(df)} 行")
                
        except Exception as e:
            return jsonify({'error': f'解析数据失败: {str(e)}'}), 400
        
        # 保存原始数据副本（用于比较变化）
        original_df = df.copy()
        
        # 准备数据详情
        df_info = {
            "row_count": len(df),
            "column_count": len(df.columns),
            "columns": list(df.columns),
            "dtypes": {col: str(df[col].dtype) for col in df.columns},
            "missing_values": df.isnull().sum().to_dict(),
            "unique_counts": {col: df[col].nunique() for col in df.columns}
        }
        
        # 准备数值型列的统计信息
        numeric_stats = {}
        for col in df.select_dtypes(include=['number']).columns:
            numeric_stats[col] = {
                "mean": float(df[col].mean()) if not pd.isna(df[col].mean()) else None,
                "median": float(df[col].median()) if not pd.isna(df[col].median()) else None,
                "min": float(df[col].min()) if not pd.isna(df[col].min()) else None,
                "max": float(df[col].max()) if not pd.isna(df[col].max()) else None,
                "std": float(df[col].std()) if not pd.isna(df[col].std()) else None
            }
        
        # 准备分类列的主要值
        categorical_stats = {}
        for col in df.select_dtypes(exclude=['number']).columns:
            # 对于每个分类列，获取前5个最常见的值
            try:
                value_counts = df[col].value_counts().head(5).to_dict()
                # 确保字典的键是字符串类型，处理Timestamp等特殊类型
                formatted_value_counts = {}
                for k, v in value_counts.items():
                    # 将键转换为字符串
                    if hasattr(k, 'strftime'):  # 如果是日期时间类型
                        key = k.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        key = str(k)
                    formatted_value_counts[key] = v
                
                categorical_stats[col] = {
                    "top_values": formatted_value_counts
                }
            except Exception as e:
                print(f"处理分类列 {col} 时出错: {str(e)}")
                categorical_stats[col] = {"top_values": {}}
        
        # 获取数据预览 - 处理特殊数据类型
        try:
            # 创建深拷贝，避免修改原始数据
            preview_df = df.head(20).copy()
            
            # 处理日期时间类型
            for col in preview_df.columns:
                if pd.api.types.is_datetime64_any_dtype(preview_df[col]):
                    preview_df[col] = preview_df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                elif preview_df[col].dtype == 'object':
                    # 尝试将其他对象转换为字符串
                    preview_df[col] = preview_df[col].astype(str)
            
            # 转换为记录格式的字典
            data_preview = preview_df.to_dict(orient='records')
        except Exception as e:
            print(f"生成数据预览时出错: {str(e)}")
            # 备用方案：逐行手动转换
            data_preview = []
            for i in range(min(20, len(df))):
                row_dict = {}
                for col in df.columns:
                    try:
                        val = df.iloc[i][col]
                        # 处理特殊类型
                        if hasattr(val, 'strftime'):  # 日期时间类型
                            row_dict[col] = val.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(val, (int, float, bool, str)) or val is None:
                            row_dict[col] = val
                        else:
                            row_dict[col] = str(val)
                    except:
                        row_dict[col] = str(val)
                data_preview.append(row_dict)
        
        # 使用阿里云大模型API处理
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {API_KEY}"
        }
        
        # 构造提示词，让AI直接返回处理后的数据
        prompt = f"""
我需要你帮我处理一个Excel数据表格，根据用户指令修改数据。

以下是数据的详细信息:
- 行数: {df_info['row_count']}
- 列数: {df_info['column_count']}
- 列名: {', '.join(df_info['columns'])}

各列的数据类型:
{json.dumps(df_info['dtypes'], indent=2)}

部分列的缺失值情况:
{json.dumps({k: v for k, v in df_info['missing_values'].items() if v > 0}, indent=2) if any(df_info['missing_values'].values()) else "无缺失值"}

数值型列的统计信息:
{json.dumps(numeric_stats, indent=2) if numeric_stats else "无数值型列"}

分类列的主要值:
{json.dumps(categorical_stats, indent=2) if categorical_stats else "无分类列"}

数据预览(前20行):
{json.dumps(data_preview, indent=2)}

用户指令是: {instruction}

请直接处理数据并返回修改后的结果。你的回复必须严格按照以下格式:

回复的第一部分是:
=== AI处理说明 ===
在这里描述你的处理思路和采取的具体步骤，以及数据发生了哪些变化。

回复的第二部分是:
=== 处理后的数据 ===
在这里仅放置一个JSON数组，格式为:
[
  {{"列名1": "值1", "列名2": "值2", ...}},
  {{"列名1": "值1", "列名2": "值2", ...}}
]

不要添加任何其他说明文字、注释或Markdown标记(如```json)。JSON数组必须能被Python的json.loads()函数直接解析。

处理数据时请注意:
1. 数据需要按照用户指令进行有意义的转换，不要简单地复制原数据
2. 如果需要排序、筛选、计算新列等，请按照用户的需求进行处理
3. 确保处理中文内容时正确
4. 所有日期类型数据必须转换为字符串格式，如"2023-01-01"
5. 所有列名和值必须用双引号包裹，数值可以不用引号
6. 不要在JSON前后添加任何注释、标记或说明文字

这是一个简单的示例输出格式:
=== AI处理说明 ===
我按照指令处理了数据...

=== 处理后的数据 ===
[
  {{"姓名": "张三", "年龄": 30}},
  {{"姓名": "李四", "年龄": 25}}
]
"""
        
        payload = {
            "model": "qwen-max",
            "input": {
                "messages": [
                    {
                        "role": "system", 
                        "content": "你是一名专业的数据处理专家，善于理解用户需求并精确执行数据处理任务。你会直接提供处理后的数据，不要返回任何解释。"
                    },
                    {"role": "user", "content": prompt}
                ]
            },
            "parameters": {
                "temperature": 0.3,
                "max_tokens": 4000
            }
        }
        
        # 调用API获取AI处理结果
        response = requests.post(API_URL, headers=headers, json=payload)
        
        if response.status_code != 200:
            return jsonify({'error': f'AI处理失败: {response.text}'}), 500
        
        # 提取AI响应
        ai_response = response.json().get('output', {}).get('text', '')
        
        # 尝试从响应中提取JSON数据
        processed_data = None
        try:
            # 先尝试从格式化的分隔标记中提取JSON
            if "=== 处理后的数据 ===" in ai_response:
                # 找到JSON数据部分
                json_part = ai_response.split("=== 处理后的数据 ===")[1].strip()
                # 提取AI解释部分
                explanation_part = ai_response.split("=== 处理后的数据 ===")[0].strip()
                if "=== AI处理说明 ===" in explanation_part:
                    explanation_part = explanation_part.split("=== AI处理说明 ===")[1].strip()
                
                # 清除可能干扰JSON解析的文本
                json_part = re.sub(r'^```json\s*', '', json_part)
                json_part = re.sub(r'\s*```$', '', json_part)
                
                try:
                    processed_data = json.loads(json_part)
                    if isinstance(processed_data, list) and len(processed_data) > 0:
                        # 成功解析，保存解释部分
                        ai_response = explanation_part
                except Exception as e:
                    print(f"从分隔标记中提取JSON失败: {str(e)}")
            
            # 如果上面的方法失败，使用正则表达式查找JSON数组
            if not processed_data:
                # 查找JSON数据块
                json_pattern = r'\[\s*\{.*\}\s*\]'  # 匹配JSON数组 [{...}, {...}]
                json_matches = re.findall(json_pattern, ai_response, re.DOTALL)
                
                if json_matches:
                    for json_str in json_matches:
                        try:
                            processed_data = json.loads(json_str)
                            if isinstance(processed_data, list) and len(processed_data) > 0:
                                # 从AI响应中删除JSON部分，只保留解释
                                ai_response = ai_response.replace(json_str, '').strip()
                                break
                        except:
                            continue
            
            # 如果仍然没找到有效JSON，尝试查找Markdown格式的JSON
            if not processed_data:
                json_pattern = r'```json\s*([\s\S]*?)\s*```'
                json_matches = re.findall(json_pattern, ai_response, re.DOTALL)
                
                if json_matches:
                    for json_str in json_matches:
                        try:
                            processed_data = json.loads(json_str)
                            if isinstance(processed_data, list) and len(processed_data) > 0:
                                # 从AI响应中删除JSON部分，只保留解释
                                full_match = f"```json{json_str}```"
                                ai_response = ai_response.replace(full_match, '').strip()
                                break
                        except:
                            continue
            
            # 如果还是没找到，尝试直接解析整个响应
            if not processed_data:
                try:
                    # 最后尝试直接解析整个响应
                    processed_data = json.loads(ai_response)
                    # 如果成功，清空解释部分，因为整个响应都是JSON
                    ai_response = ""
                except:
                    pass
                    
            # 验证处理后的数据是否有效
            if not processed_data or not isinstance(processed_data, list) or len(processed_data) == 0:
                # 如果未能提取有效数据，尝试获取解释部分
                explanation = ai_response.split('[')[0].strip() if '[' in ai_response else ai_response
                return jsonify({
                    'error': '无法从AI响应中提取有效的数据',
                    'ai_explanation': explanation
                }), 500
                
        except Exception as e:
            print(f"提取处理数据失败: {str(e)}")
            print(f"AI响应: {ai_response[:500]}...")
            return jsonify({'error': f'解析AI返回的数据失败: {str(e)}'}), 500
        
        # 将AI处理的数据转换回DataFrame
        try:
            # 处理数据以确保没有不兼容的类型
            clean_data = []
            for item in processed_data:
                clean_item = {}
                for k, v in item.items():
                    # 处理可能的非标准类型
                    if hasattr(v, 'strftime'):  # 如果是日期时间类型
                        clean_item[k] = v.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(v, (int, float, bool, str)) or v is None:
                        clean_item[k] = v
                    else:
                        clean_item[k] = str(v)
                clean_data.append(clean_item)
            
            processed_df = pd.DataFrame(clean_data)
            
            # 如果返回的数据太大，只保留最多前1000行
            if len(processed_df) > 1000:
                processed_df = processed_df.head(1000)
                print(f"AI返回的数据太大，截取前1000行")
            
            # 保存处理后的文件
            ai_processed_filename = f'ai_processed_{str(uuid.uuid4())}.xlsx'
            ai_processed_filepath = os.path.join(PROCESSED_FOLDER, ai_processed_filename)
            processed_df.to_excel(ai_processed_filepath, index=False)
            
            # 生成预览数据 - 处理特殊类型，确保JSON序列化不会出错
            preview_df = processed_df.head(10).copy()
            for col in preview_df.columns:
                if pd.api.types.is_datetime64_any_dtype(preview_df[col]):
                    preview_df[col] = preview_df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                elif preview_df[col].dtype == 'object':
                    # 尝试将其他对象转换为字符串
                    preview_df[col] = preview_df[col].astype(str)
            
            new_preview_data = preview_df.to_json(orient='split')
            
            # 提取AI解释（第一部分文本，不包含JSON）
            explanation = ai_response.split('[')[0].strip() if '[' in ai_response else ""
            
            # 生成结果摘要
            changed_cols = [col for col in processed_df.columns if col not in original_df.columns]
            removed_cols = [col for col in original_df.columns if col not in processed_df.columns]
            
            # 识别修改的列（值发生变化但列名未变）
            modified_cols = []
            common_cols = set(original_df.columns).intersection(set(processed_df.columns))
            for col in common_cols:
                if col not in changed_cols and col not in removed_cols:
                    # 只检查前100行，提高性能
                    original_sample = original_df[col].head(100) if len(original_df) > 0 else []
                    processed_sample = processed_df[col].head(100) if len(processed_df) > 0 else []
                    
                    try:
                        if len(original_sample) != len(processed_sample) or not original_sample.equals(processed_sample):
                            modified_cols.append(col)
                    except:
                        # 如果比较失败（可能是数据类型不同），也认为列已修改
                        modified_cols.append(col)
            
            summary = {
                "new_columns": changed_cols,
                "removed_columns": removed_cols,
                "modified_columns": modified_cols[:10],  # 限制显示的数量
                "total_rows": len(processed_df),
                "total_cols": len(processed_df.columns),
                "instruction": instruction,
                "has_changes": len(changed_cols) > 0 or len(removed_cols) > 0 or len(modified_cols) > 0 or len(processed_df) != len(original_df)
            }
            
            return jsonify({
                'success': True,
                'preview': new_preview_data,
                'summary': summary,
                'download_filename': ai_processed_filename,
                'ai_explanation': explanation
            })
            
        except Exception as e:
            error_info = str(e)
            traceback_info = traceback.format_exc()
            print(f"处理AI返回数据时出错: {error_info}\n{traceback_info}")
            
            # 返回错误信息
            return jsonify({
                'error': f'处理AI返回的数据失败: {str(e)}',
                'traceback': traceback_info
            }), 500
            
    except Exception as e:
        return jsonify({'error': f'处理请求失败: {str(e)}', 'traceback': traceback.format_exc()}), 500

@app.route('/')
def index():
    # 返回前端 HTML 页面
    return render_template('index.html')
@app.route('/api/download_analysis/<filename>')
def download_analysis_file(filename):
    filepath = os.path.join(ANALYSIS_FOLDER, filename)
    try:
        if os.path.exists(filepath):
            try:
                # 提取数据分析信息
                df_preview = pd.read_excel(filepath, sheet_name='数据预览')
                df_stats = pd.read_excel(filepath, sheet_name='基本统计')
                df_analysis = pd.read_excel(filepath, sheet_name='AI分析')
                analysis_report = df_analysis.iloc[0, 0] if not df_analysis.empty else "无法提取分析报告内容"
                
                # 获取分析ID从文件名
                analysis_id = filename.replace('analysis_', '').replace('.xlsx', '')
                
                # 检查可视化图像是否存在
                img_dir = os.path.join(ANALYSIS_FOLDER, f'images_{analysis_id}')
                images = []
                if os.path.exists(img_dir):
                    for i in range(1, 6):  # 最多5张图
                        img_path = os.path.join(img_dir, f'plot_{i}.png')
                        if os.path.exists(img_path):
                            with open(img_path, 'rb') as img_file:
                                img_data = base64.b64encode(img_file.read()).decode('utf-8')
                                images.append(img_data)
                
                # 创建HTML报告
                html_filename = f'report_{analysis_id}.html'
                html_filepath = os.path.join(ANALYSIS_FOLDER, html_filename)
                
                # 基本统计数据转换为HTML表格
                stats_html = df_stats.to_html(index=False)
                
                # 数据预览转换为HTML表格
                preview_html = df_preview.head(10).to_html(index=False)
                
                # 转换Markdown格式的分析报告到HTML
                analysis_html = analysis_report.replace("**", "<strong>").replace("**", "</strong>")
                analysis_html = analysis_html.replace("\n\n", "<br><br>")
                analysis_html = analysis_html.replace("\n- ", "<br>• ")
                
                # 生成完整HTML报告
                html_content = f"""
                <!DOCTYPE html>
                <html lang="zh-CN">
                <head>
                    <meta charset="UTF-8">
                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                    <title>数据分析报告</title>
                    <style>
                        body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif; margin: 0; padding: 20px; line-height: 1.6; color: #333; }}
                        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                        h1, h2, h3, h4 {{ color: #2c3e50; }}
                        h1 {{ font-size: 28px; margin-top: 0; text-align: center; color: #73BF00; }}
                        h2 {{ font-size: 22px; margin-top: 30px; border-bottom: 2px solid #73BF00; padding-bottom: 10px; }}
                        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                        table, th, td {{ border: 1px solid #ddd; }}
                        th {{ background-color: #f2f2f2; text-align: left; padding: 12px; }}
                        td {{ padding: 12px; }}
                        .image-gallery {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(400px, 1fr)); gap: 20px; margin: 20px 0; }}
                        .image-container {{ background: white; padding: 10px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
                        img {{ max-width: 100%; height: auto; border-radius: 4px; }}
                        .section {{ margin-bottom: 40px; }}
                        .footer {{ text-align: center; margin-top: 40px; font-size: 14px; color: #7f8c8d; }}
                        strong {{ font-weight: bold; }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <h1>数据分析报告</h1>
                        
                        <div class="section">
                            <h2>1. 基本统计信息</h2>
                            {stats_html}
                        </div>
                        
                        <div class="section">
                            <h2>2. 数据预览</h2>
                            {preview_html}
                        </div>
                        
                        <div class="section">
                            <h2>3. AI分析报告</h2>
                            <div>{analysis_html}</div>
                        </div>
                """
                
                # 添加可视化图表部分
                if images:
                    html_content += """
                        <div class="section">
                            <h2>4. 数据可视化</h2>
                            <div class="image-gallery">
                    """
                    
                    for i, img_data in enumerate(images):
                        html_content += f"""
                                <div class="image-container">
                                    <img src="data:image/png;base64,{img_data}" alt="图表 {i+1}">
                                    <p>图表 {i+1}</p>
                                </div>
                        """
                    
                    html_content += """
                            </div>
                        </div>
                    """
                
                # 添加页脚
                html_content += f"""
                        <div class="footer">
                            <p>报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                        </div>
                    </div>
                </body>
                </html>
                """
                
                # 写入HTML文件
                with open(html_filepath, 'w', encoding='utf-8') as f:
                    f.write(html_content)
                
                # 返回HTML文件
                return send_file(html_filepath, as_attachment=True, download_name="数据分析报告.html")
                
            except Exception as e:
                print(f"创建报告失败: {str(e)}")
                # 如果创建报告失败，则返回原始Excel文件
                return send_file(filepath, as_attachment=True)
        else:
            return jsonify({
                'error': '分析文件未找到',
                'filepath': filepath
            }), 404
    except Exception as e:
        return jsonify({
            'error': f'下载分析文件失败: {str(e)}',
            'filepath': filepath
        }), 500

@app.route('/api/data_analysis', methods=['POST'])
def data_analysis():
    if 'file' not in request.files:
        return jsonify({'error': '没有文件上传'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls') or file.filename.endswith('.csv')):
        temp_filepath = None
        try:
            filename = secure_filename(file.filename)
            temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(temp_filepath)

            # 读取Excel数据
            if file.filename.endswith('.csv'):
                df = pd.read_csv(temp_filepath)
            else:
                df = pd.read_excel(temp_filepath)

            # 基本数据统计
            basic_stats = {
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": df.columns.tolist(),
                "missing_values": df.isnull().sum().to_dict(),
                "data_types": {col: str(df[col].dtype) for col in df.columns}
            }

            # 数值型列的统计信息
            numeric_stats = {}
            for col in df.select_dtypes(include=['number']).columns:
                numeric_stats[col] = {
                    "mean": float(df[col].mean()) if not pd.isna(df[col].mean()) else None,
                    "median": float(df[col].median()) if not pd.isna(df[col].median()) else None,
                    "std": float(df[col].std()) if not pd.isna(df[col].std()) else None,
                    "min": float(df[col].min()) if not pd.isna(df[col].min()) else None,
                    "max": float(df[col].max()) if not pd.isna(df[col].max()) else None
                }

            # 分类列的统计信息
            categorical_stats = {}
            for col in df.select_dtypes(include=['object', 'category']).columns:
                value_counts = df[col].value_counts().head(10).to_dict()  # 只取前10个类别
                categorical_stats[col] = {
                    "unique_values": df[col].nunique(),
                    "top_categories": value_counts
                }

            # 生成数据预览
            preview_data = df.head(10).to_json(orient='split')

            # 生成分析ID
            analysis_id = str(uuid.uuid4())

            # 使用AI生成深度分析报告
            analysis_report = generate_ai_analysis(df, basic_stats, numeric_stats, categorical_stats)
            
            # 使用AI生成可视化代码
            visualization_code = generate_ai_visualization_code(df, numeric_stats, categorical_stats)
            
            # 执行可视化代码并获取图像
            visualization_result = execute_visualization_code(visualization_code, df, analysis_id)
            
            # 保存分析结果
            result_filename = f'analysis_{analysis_id}.xlsx'
            result_filepath = os.path.join(ANALYSIS_FOLDER, result_filename)
            
            # 创建带有分析结果的Excel文件
            with pd.ExcelWriter(result_filepath) as writer:
                df.head(100).to_excel(writer, sheet_name='数据预览', index=False)
                
                # 创建统计信息表
                stats_df = pd.DataFrame({
                    "统计项": ["总行数", "总列数", "数值型列数", "类别型列数"],
                    "值": [
                        basic_stats["rows"], 
                        basic_stats["columns"],
                        len(numeric_stats),
                        len(categorical_stats)
                    ]
                })
                stats_df.to_excel(writer, sheet_name='基本统计', index=False)
                
                # 将AI分析报告保存为单独的表格
                pd.DataFrame({"AI分析报告": [analysis_report]}).to_excel(writer, sheet_name='AI分析', index=False)
                
                # 将可视化代码保存为单独的表格
                pd.DataFrame({"数据可视化代码": [visualization_code]}).to_excel(writer, sheet_name='可视化代码', index=False)

            # 清理临时文件
            os.remove(temp_filepath)

            # 构建响应数据
            response_data = {
                'success': True,
                'basic_stats': basic_stats,
                'numeric_stats': numeric_stats,
                'categorical_stats': categorical_stats,
                'preview': preview_data,
                'analysis_report': analysis_report,
                'download_filename': result_filename,
                'analysis_id': analysis_id
            }
            
            # 添加可视化结果
            if visualization_result['success']:
                response_data['visualization'] = {
                    'success': True,
                    'images': visualization_result['image_b64_list'],
                    'code': visualization_result['code']
                }
            else:
                response_data['visualization'] = {
                    'success': False,
                    'error': visualization_result['error'],
                    'code': visualization_result['code']
                }

            return jsonify(response_data)

        except Exception as e:
            if temp_filepath and os.path.exists(temp_filepath):
                os.remove(temp_filepath)
            return jsonify({'error': f'数据分析时发生错误: {str(e)}', 'traceback': traceback.format_exc()}), 500
    else:
        return jsonify({'error': '不支持的文件类型，请上传 .xls, .xlsx 或 .csv 文件'}), 400

def generate_ai_analysis(df, basic_stats, numeric_stats, categorical_stats):
    """使用AI生成数据分析报告"""
    try:
        # 准备数据摘要
        data_summary = f"""
数据概览:
- 总行数: {basic_stats['rows']}
- 总列数: {basic_stats['columns']}
- 列名: {', '.join(basic_stats['column_names'][:10])}{'...' if len(basic_stats['column_names']) > 10 else ''}

数值型列统计:
"""
        for col, stats in list(numeric_stats.items())[:5]:  # 只取前5个数值列
            data_summary += f"- {col}: 均值={stats['mean']}, 中位数={stats['median']}, 最小值={stats['min']}, 最大值={stats['max']}, 标准差={stats['std']}\n"
        
        data_summary += "\n分类型列统计:\n"
        for col, stats in list(categorical_stats.items())[:5]:  # 只取前5个分类列
            data_summary += f"- {col}: 唯一值数量={stats['unique_values']}, 主要类别={list(stats['top_categories'].keys())[:3]}\n"

        # 数据预览（前5行，只显示部分列）
        preview_cols = basic_stats['column_names'][:10]  # 最多10列
        data_sample = df[preview_cols].head(5).to_string()
        data_summary += f"\n数据样例（前5行）:\n{data_sample}\n"

        # 构造AI请求
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {API_KEY}"
        }

        prompt = f"""
作为一名专业的数据分析师，请基于以下数据信息，提供一份深入的数据分析报告。
这份报告至少包含：

1. 数据概述（数据规模、类型和基本结构）
2. 关键洞察和发现（数据的主要特征和模式）
3. 潜在问题点（数据质量问题、异常值等）
4. 具体建议（如何利用此数据进行决策或进一步分析）
5. 本数据的价值以及更远的展望

以下是数据的基本信息：
{data_summary}

请提供一份专业、有深度且有洞察力的分析报告，尽量挖掘数据中的价值和模式。报告应该是结构化的，便于阅读，并包含实际可行的建议。
"""

        payload = {
            "model": "qwen-max",
            "input": {
                "messages": [
                    {
                        "role": "system", 
                        "content": "你是一名专业的数据分析师，擅长从数据中发现洞察和价值，并提供专业的分析报告。"
                    },
                    {"role": "user", "content": prompt}
                ]
            },
            "parameters": {
                "temperature": 0.5,
                "max_tokens": 2000
            }
        }

        # 调用API获取分析报告
        response = requests.post(API_URL, headers=headers, json=payload)
        
        if response.status_code == 200:
            ai_response = response.json().get('output', {}).get('text', '')
            return ai_response
        else:
            return f"AI分析生成失败: HTTP {response.status_code} - {response.text}"

    except Exception as e:
        return f"生成AI分析报告时出错: {str(e)}"

def generate_ai_visualization_code(df, numeric_stats, categorical_stats):
    """使用AI生成数据可视化代码"""
    try:
        # 准备数据摘要
        data_description = f"数据总行数: {len(df)}\n数据总列数: {len(df.columns)}\n"
        
        # 详细描述数据结构
        data_description += "列名及其数据类型:\n"
        for col in df.columns:
            data_description += f"- {col}: {df[col].dtype}\n"
        
        # 描述缺失值情况
        missing_values = df.isnull().sum()
        if missing_values.sum() > 0:
            data_description += "\n缺失值情况:\n"
            for col, count in missing_values[missing_values > 0].items():
                percent = (count / len(df)) * 100
                data_description += f"- {col}: {count}个缺失值 ({percent:.2f}%)\n"
        
        # 数值列信息
        numeric_cols = list(numeric_stats.keys())
        if numeric_cols:
            data_description += f"\n数值型列: {', '.join(numeric_cols)}\n"
            
            # 添加一些统计信息示例
            for col in numeric_cols[:3]:
                stats = numeric_stats[col]
                data_description += f"- {col} - 均值: {stats['mean']}, 中位数: {stats['median']}, 最小值: {stats['min']}, 最大值: {stats['max']}, 标准差: {stats['std']}\n"
        
        # 分类列信息
        categorical_cols = list(categorical_stats.keys())
        if categorical_cols:
            data_description += f"\n分类型列: {', '.join(categorical_cols)}\n"
            
            # 添加一些分类信息示例
            for col in categorical_cols[:3]:
                stats = categorical_stats[col]
                data_description += f"- {col} - 唯一值数量: {stats['unique_values']}, 主要类别: {list(stats['top_categories'].keys())[:3]}\n"
        
        # 提供一些数据示例
        data_description += f"\n数据示例（前3行）:\n{df.head(3).to_string()}\n"
        
        # 构造AI请求
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {API_KEY}"
        }

        prompt = f"""
请为以下数据生成5种不同类型的数据可视化Python代码。

数据描述:
{data_description}

要求:
1. 生成5种不同的可视化代码，每种代码应该能展示数据的不同方面
2. 使用matplotlib, seaborn, numpy库
3. 设置合适的中文标题、轴标签和图例
4. 每个图表都应该提供有价值的数据洞察
5. 确保代码能处理缺失值（使用dropna()或fillna()方法）
6. 每个代码块使用函数形式，函数名分别为plot1, plot2, plot3, plot4, plot5
7. 代码必须包含合适的注释，解释每个图表的目的和洞察
8. 图表类型可以包括：条形图、散点图、箱线图、热力图、饼图、折线图等
9. 适当调整图表样式，确保美观
10. 函数参数为df (pandas DataFrame)
11. 务必确保正确的代码缩进
12. 在代码中加入异常处理，确保即使某些列不存在也不会报错
13. 确保代码可以在不同数据集上运行

重要：针对中文显示问题，请在每个绘图函数的开头添加以下代码来确保中文正确显示：
```python
# 设置中文显示
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'PingFang SC', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False
```

同时，对于包含中文的标题、标签和图例，请使用以下方式设置：
```python
plt.title("中文标题", fontproperties='SimHei')
plt.xlabel("中文横轴", fontproperties='SimHei')
plt.ylabel("中文纵轴", fontproperties='SimHei')
```

请确保使用上述方法处理所有中文文本，这对于正确显示图表中的中文文字至关重要。

仅返回可执行的Python代码，不要有解释性文字。
"""

        payload = {
            "model": "qwen-max",
            "input": {
                "messages": [
                    {
                        "role": "system", 
                        "content": "你是一名数据可视化专家，精通使用Python进行数据可视化编程。你的代码必须是可执行的，没有语法错误，并且能够处理各种不同的数据集。"
                    },
                    {"role": "user", "content": prompt}
                ]
            },
            "parameters": {
                "temperature": 0.3,
                "max_tokens": 3000
            }
        }

        # 调用API获取可视化代码
        response = requests.post(API_URL, headers=headers, json=payload)
        
        if response.status_code == 200:
            ai_response = response.json().get('output', {}).get('text', '')
            
            # 提取代码块
            code_pattern = r"```python\s*(.*?)\s*```"
            code_match = re.search(code_pattern, ai_response, re.DOTALL)
            
            if code_match:
                return code_match.group(1).strip()
            else:
                # 如果没有代码块标记，尝试直接使用整个响应
                return ai_response.strip()
        else:
            return f"生成可视化代码失败: HTTP {response.status_code}"

    except Exception as e:
        return f"生成可视化代码时出错: {str(e)}"

def execute_visualization_code(code, df, analysis_id):
    """执行AI生成的可视化代码并保存图像"""
    image_paths = []
    image_b64_list = []
    
    try:
        # 创建图片保存目录
        img_dir = os.path.join(ANALYSIS_FOLDER, f'images_{analysis_id}')
        if not os.path.exists(img_dir):
            os.makedirs(img_dir)
            
        # 在代码执行前重新配置中文字体设置
        try:
            # 查找系统中可用的中文字体
            font_found = False
            
            # macOS 中文字体路径
            mac_fonts = [
                '/System/Library/Fonts/PingFang.ttc',
                '/System/Library/Fonts/STHeiti Light.ttc',
                '/System/Library/Fonts/STHeiti Medium.ttc',
                '/Library/Fonts/Microsoft/SimHei.ttf',
                '/Library/Fonts/Microsoft/Microsoft YaHei.ttf'
            ]
            
            # 检查并使用第一个找到的字体
            for font_path in mac_fonts:
                if os.path.exists(font_path):
                    font = FontProperties(fname=font_path)
                    plt.rcParams['font.family'] = ['sans-serif']
                    plt.rcParams['font.sans-serif'] = ['SimHei', 'PingFang SC', 'Microsoft YaHei', 'STHeiti', 'Arial Unicode MS', 'sans-serif']
                    matplotlib.rcParams['font.sans-serif'] = ['SimHei', 'PingFang SC', 'Microsoft YaHei', 'STHeiti', 'Arial Unicode MS', 'sans-serif']
                    print(f"使用字体: {font_path}")
                    font_found = True
                    break
            
            # 如果没有找到合适的字体，使用系统默认配置
            if not font_found:
                print("未找到合适的中文字体，使用系统默认字体")
                plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'sans-serif']
                
            # 确保负号正确显示
            plt.rcParams['axes.unicode_minus'] = False
            
            # 设置字号
            plt.rcParams['font.size'] = 12
            plt.rcParams['axes.titlesize'] = 14
            plt.rcParams['axes.labelsize'] = 12
            
        except Exception as font_e:
            print(f"字体配置错误: {str(font_e)}")
        
        # 定义一个捕获matplotlib图表的函数
        def save_current_figure(index):
            # 添加标题和标签的中文修复
            try:
                for ax in plt.gcf().get_axes():
                    # 获取当前标题和标签
                    title = ax.get_title()
                    xlabel = ax.get_xlabel()
                    ylabel = ax.get_ylabel()
                    
                    # 如果包含中文，添加font属性
                    if title and any(ord(c) > 127 for c in title):
                        ax.set_title(title, fontproperties=font if 'font' in locals() else None)
                    if xlabel and any(ord(c) > 127 for c in xlabel):
                        ax.set_xlabel(xlabel, fontproperties=font if 'font' in locals() else None)
                    if ylabel and any(ord(c) > 127 for c in ylabel):
                        ax.set_ylabel(ylabel, fontproperties=font if 'font' in locals() else None)
                    
                    # 修复刻度标签
                    for label in ax.get_xticklabels() + ax.get_yticklabels():
                        if any(ord(c) > 127 for c in label.get_text()):
                            label.set_fontproperties(font if 'font' in locals() else None)
            except Exception as e:
                print(f"修复图表中文显示错误: {str(e)}")
            
            img_path = os.path.join(img_dir, f'plot_{index}.png')
            plt.savefig(img_path, dpi=100, bbox_inches='tight')
            
            # 转换为base64以便在网页显示
            buffered = io.BytesIO()
            plt.savefig(buffered, format="png", dpi=100, bbox_inches='tight')
            img_str = base64.b64encode(buffered.getvalue()).decode()
            
            plt.close()  # 关闭图表释放内存
            return img_path, img_str
        
        # 检查并修复常见的代码问题
        # 修复可能的缩进问题
        lines = code.split('\n')
        fixed_lines = []
        for i, line in enumerate(lines):
            # 忽略空行和注释行
            if not line.strip() or line.strip().startswith('#'):
                fixed_lines.append(line)
                continue
                
            # 检查函数定义和函数体的缩进关系
            if line.strip().startswith('def '):
                # 这是一个函数定义，记录其缩进
                indent_level = len(line) - len(line.lstrip())
                expected_indent = indent_level + 4  # 预期的函数体缩进
                fixed_lines.append(line)
            elif i > 0 and lines[i-1].strip().startswith('def '):
                # 这是函数定义后的第一行，应该有4个空格的缩进
                current_indent = len(line) - len(line.lstrip())
                if current_indent < 4:
                    # 修复缩进
                    fixed_lines.append(' ' * 4 + line.lstrip())
                else:
                    fixed_lines.append(line)
            else:
                fixed_lines.append(line)
        
        fixed_code = '\n'.join(fixed_lines)
        
        # 检查数据列是否存在
        # 提取代码中使用的列名
        column_pattern = r"df\['([^']+)'\]"
        used_columns = re.findall(column_pattern, fixed_code)
        
        # 检查这些列是否存在于数据框中
        for col in set(used_columns):
            if col not in df.columns:
                # 列不存在，修改代码使用可用的列
                if df.select_dtypes(include=['number']).columns.size > 0:
                    # 用第一个数值列替换
                    numeric_col = df.select_dtypes(include=['number']).columns[0]
                    fixed_code = fixed_code.replace(f"df['{col}']", f"df['{numeric_col}']")
                elif df.columns.size > 0:
                    # 用第一个列替换
                    first_col = df.columns[0]
                    fixed_code = fixed_code.replace(f"df['{col}']", f"df['{first_col}']")
        
        # 准备执行环境
        local_vars = {
            'df': df,
            'plt': plt,
            'sns': sns,
            'np': np,
            'pd': pd
        }
        
        # 执行修复后的代码
        exec(fixed_code, local_vars)
        
        # 检查并执行生成的函数
        plot_functions = []
        for i in range(1, 6):
            func_name = f'plot{i}'
            if func_name in local_vars and callable(local_vars[func_name]):
                plot_functions.append((i, local_vars[func_name]))
        
        # 在生成的代码中注入中文支持
        # 检查是否需要修改生成的绘图代码以支持中文显示
        modified_code = fixed_code
        
        # 尝试修改可能存在的中文问题
        try:
            # 在绘图函数中添加中文字体支持
            for i in range(1, 6):
                # 查找plot函数定义
                func_pattern = r"def plot{}".format(i)
                if re.search(func_pattern, modified_code):
                    # 找到函数定义所在行
                    lines = modified_code.split('\n')
                    for j, line in enumerate(lines):
                        if re.search(func_pattern, line):
                            # 查找函数体的第一行
                            indent = len(line) - len(line.lstrip())
                            # 在函数体的开始添加中文字体支持代码
                            font_support_code = ' ' * (indent + 4) + "# 确保中文显示正常\n"
                            font_support_code += ' ' * (indent + 4) + "plt.rcParams['font.sans-serif'] = ['SimHei', 'PingFang SC', 'Microsoft YaHei']\n"
                            font_support_code += ' ' * (indent + 4) + "plt.rcParams['axes.unicode_minus'] = False\n"
                            
                            # 找到函数体的第一行
                            k = j + 1
                            while k < len(lines) and (not lines[k].strip() or lines[k].lstrip().startswith('#')):
                                k += 1
                            
                            if k < len(lines):
                                # 插入字体支持代码
                                lines.insert(k, font_support_code)
                                
                            break
                    
                    modified_code = '\n'.join(lines)
        except Exception as e:
            print(f"修改绘图代码以支持中文时出错: {str(e)}")
        
        # 重新执行修改后的代码
        try:
            exec(modified_code, local_vars)
        except Exception as e:
            print(f"执行修改后的代码出错，将继续使用原始代码: {str(e)}")
        
        # 执行每个绘图函数并保存结果
        for idx, func in plot_functions:
            try:
                plt.figure(figsize=(10, 6))
                func(df)  # 调用绘图函数
                
                # 设置全局字体
                plt.rcParams['font.sans-serif'] = ['SimHei', 'PingFang SC', 'Microsoft YaHei', 'STHeiti', 'Arial Unicode MS']
                plt.rcParams['axes.unicode_minus'] = False
                
                img_path, img_b64 = save_current_figure(idx)
                image_paths.append(img_path)
                image_b64_list.append({
                    'index': idx,
                    'base64': img_b64
                })
            except Exception as e:
                print(f"绘制图表 {idx} 时出错: {str(e)}")
                traceback.print_exc()  # 打印详细的堆栈信息
                # 继续处理下一个图表，不中断整个流程
        
        # 如果没有成功创建任何图表，返回失败信息
        if not image_paths:
            return {
                'success': False,
                'error': "无法生成任何可视化图表",
                'code': fixed_code
            }
        
        return {
            'success': True,
            'image_paths': image_paths,
            'image_b64_list': image_b64_list,
            'code': fixed_code  # 返回修复后的代码
        }
        
    except Exception as e:
        error_traceback = traceback.format_exc()
        print(f"可视化代码执行错误: {str(e)}\n{error_traceback}")
        return {
            'success': False,
            'error': str(e),
            'traceback': error_traceback,
            'code': code
        }

@app.route('/api/analysis_image/<analysis_id>/<image_name>')
def get_analysis_image(analysis_id, image_name):
    """获取数据分析生成的图片"""
    img_dir = os.path.join(ANALYSIS_FOLDER, f'images_{analysis_id}')
    img_path = os.path.join(img_dir, image_name)
    
    try:
        if os.path.exists(img_path):
            return send_file(img_path, mimetype='image/png')
        else:
            return jsonify({'error': '图片未找到'}), 404
    except Exception as e:
        return jsonify({'error': f'获取图片失败: {str(e)}'}), 500

if __name__ == '__main__':
    print(f"Excel处理文件夹路径: {PROCESSED_FOLDER}")
    app.run(debug=True, host='0.0.0.0', port=8080)
